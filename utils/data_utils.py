from __future__ import annotations

import json
import re
from pathlib import Path
from typing import List, Tuple

import numpy as np
import pandas as pd
from sklearn.preprocessing import MinMaxScaler, StandardScaler


def set_seed(seed: int) -> None:
    import random
    import torch

    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def generate_synthetic_multidomain_data(num_steps: int, random_seed: int) -> pd.DataFrame:
    rng = np.random.default_rng(random_seed)
    idx = pd.date_range("2023-01-01", periods=num_steps, freq="h")
    t = np.arange(num_steps)

    daily = np.sin(2 * np.pi * t / 24)
    weekly = np.sin(2 * np.pi * t / (24 * 7))
    annual = np.sin(2 * np.pi * t / (24 * 365))

    temperature = 24 + 8 * daily + 3 * annual + rng.normal(0, 1.2, num_steps)
    humidity = 55 - 12 * daily + 6 * weekly + rng.normal(0, 2.2, num_steps)
    wind_speed = 8 + 1.5 * np.cos(2 * np.pi * t / 24) + rng.normal(0, 0.7, num_steps)

    commute_profile = np.maximum(0, np.sin(2 * np.pi * (t - 6) / 24)) + np.maximum(0, np.sin(2 * np.pi * (t - 16) / 24))
    traffic = 120 + 42 * commute_profile + 18 * weekly + 0.35 * humidity + rng.normal(0, 4.5, num_steps)

    electricity = (
        210
        + 2.8 * temperature
        + 0.42 * traffic
        + 20 * np.maximum(0, daily)
        + 10 * weekly
        + rng.normal(0, 5.5, num_steps)
    )

    pollution = (
        40
        + 0.32 * traffic
        + 0.09 * electricity
        - 1.7 * wind_speed
        + 0.18 * humidity
        + 4.5 * np.maximum(0, daily)
        + rng.normal(0, 3.0, num_steps)
    )

    drift_start = int(num_steps * 0.72)
    pollution[drift_start:] += 0.13 * electricity[drift_start:] + 0.12 * temperature[drift_start:]
    traffic[drift_start:] += 14 * np.sin(2 * np.pi * t[drift_start:] / 12)
    electricity[drift_start:] += 18 * np.maximum(0, np.sin(2 * np.pi * (t[drift_start:] - 4) / 24))

    df = pd.DataFrame(
        {
            "timestamp": idx,
            "traffic_flow": traffic,
            "aqi": pollution,
            "electricity_demand": electricity,
            "temperature": temperature,
            "humidity": humidity,
        }
    )
    return df


def _needs_rebuild(output_path: Path, source_paths: list[Path]) -> bool:
    if not output_path.exists():
        return True
    output_mtime = output_path.stat().st_mtime
    return any(path.exists() and path.stat().st_mtime > output_mtime for path in source_paths)


def _to_numeric_series(series: pd.Series) -> pd.Series:
    cleaned = series.astype(str).str.replace(",", ".", regex=False)
    numeric = pd.to_numeric(cleaned, errors="coerce")
    return numeric.replace(-200, np.nan)


def _expand_daily_traffic_to_hourly(traffic_daily: pd.DataFrame) -> pd.DataFrame:
    hourly_profile = np.array(
        [
            0.022,
            0.018,
            0.015,
            0.015,
            0.020,
            0.032,
            0.051,
            0.064,
            0.069,
            0.056,
            0.047,
            0.043,
            0.041,
            0.042,
            0.045,
            0.051,
            0.063,
            0.077,
            0.083,
            0.071,
            0.056,
            0.044,
            0.032,
            0.027,
        ],
        dtype=float,
    )
    hourly_profile /= hourly_profile.sum()

    frames = []
    for _, row in traffic_daily.iterrows():
        timestamps = pd.date_range(row["timestamp"], periods=24, freq="h")
        day_type_factor = np.where(timestamps.dayofweek >= 5, 0.88, 1.0)
        profile = hourly_profile * day_type_factor
        profile /= profile.sum()
        frames.append(
            pd.DataFrame(
                {
                    "timestamp": timestamps,
                    "traffic_flow": row["traffic_flow"] * profile,
                    "traffic_speed": row["traffic_speed"],
                    "traffic_congestion": row["traffic_congestion"],
                    "traffic_incidents": row["traffic_incidents"] * profile,
                    "traffic_environmental_impact": row["traffic_environmental_impact"] * profile,
                }
            )
        )

    return pd.concat(frames, ignore_index=True)


def _resample_sequence(values: np.ndarray, target_len: int) -> np.ndarray:
    if len(values) == 0:
        raise ValueError("Cannot resample an empty sequence")
    if len(values) == target_len:
        return values.astype(float)

    source_index = np.linspace(0, 1, num=len(values))
    target_index = np.linspace(0, 1, num=target_len)
    return np.interp(target_index, source_index, values.astype(float))


def _smooth_sequence(values: np.ndarray, window: int = 6) -> np.ndarray:
    series = pd.Series(np.asarray(values, dtype=float))
    return series.rolling(window=window, min_periods=1, center=True).mean().to_numpy()


def _calendar_align(
    source_df: pd.DataFrame,
    target_timestamps: pd.Series,
    value_columns: list[str],
    grouping_levels: list[tuple[str, ...]],
) -> pd.DataFrame:
    source = source_df.copy()
    target = pd.DataFrame({"timestamp": pd.to_datetime(target_timestamps)})

    for frame in (source, target):
        frame["month"] = frame["timestamp"].dt.month
        frame["day_of_week"] = frame["timestamp"].dt.dayofweek
        frame["hour"] = frame["timestamp"].dt.hour

    aligned = pd.DataFrame(index=target.index, columns=value_columns, dtype=float)
    remaining = pd.Series(True, index=target.index)

    for level in grouping_levels:
        lookup = source.groupby(list(level))[value_columns].median().reset_index()
        merged = target.loc[remaining, list(level)].merge(lookup, how="left", on=list(level))
        merged.index = target.index[remaining]
        for col in value_columns:
            aligned.loc[merged.index, col] = aligned.loc[merged.index, col].fillna(merged[col])
        remaining = aligned[value_columns].isna().any(axis=1)
        if not remaining.any():
            break

    global_fill = source[value_columns].median()
    for col in value_columns:
        aligned[col] = aligned[col].fillna(global_fill[col])
    return aligned.reset_index(drop=True)


def _blend_aligned_signals(calendar_values: np.ndarray, resampled_values: np.ndarray, calendar_weight: float) -> np.ndarray:
    blended = calendar_weight * np.asarray(calendar_values, dtype=float) + (1.0 - calendar_weight) * np.asarray(
        resampled_values,
        dtype=float,
    )
    return _smooth_sequence(blended)


def _build_electricity_baseline(
    electricity_yearly: pd.DataFrame,
    target_timestamps: pd.Series,
    traffic_flow: np.ndarray,
    temperature: np.ndarray,
    humidity: np.ndarray,
) -> np.ndarray:
    yearly = electricity_yearly[["timestamp", "Energy Input in MU"]].dropna().sort_values("timestamp").copy()
    target_index = pd.to_datetime(target_timestamps)

    ordinal_target = target_index.astype("int64").to_numpy(dtype=float)
    ordinal_source = yearly["timestamp"].astype("int64").to_numpy(dtype=float)
    yearly_signal = np.interp(ordinal_target, ordinal_source, yearly["Energy Input in MU"].to_numpy(dtype=float))

    hour = target_index.dt.hour.to_numpy()
    day_of_week = target_index.dt.dayofweek.to_numpy()
    seasonal = 1.0 + 0.04 * np.sin(2 * np.pi * target_index.dt.dayofyear.to_numpy() / 365.0)
    weekday_factor = np.where(day_of_week >= 5, 0.95, 1.0)
    intraday = 12.0 * np.maximum(0, np.sin(2 * np.pi * (hour - 8) / 24))

    traffic_component = 0.00010 * np.asarray(traffic_flow, dtype=float)
    weather_component = 2.3 * np.asarray(temperature, dtype=float) + 0.05 * np.asarray(humidity, dtype=float)
    return _smooth_sequence((0.0065 * yearly_signal * seasonal + intraday + traffic_component + weather_component) * weekday_factor)


def _minmax_scale(values: np.ndarray) -> np.ndarray:
    values = np.asarray(values, dtype=float)
    lower = np.nanmin(values)
    upper = np.nanmax(values)
    if np.isclose(lower, upper):
        return np.zeros_like(values)
    return (values - lower) / (upper - lower)


def _compute_traffic_sensitive_aqi(
    baseline_aqi: np.ndarray,
    traffic_flow: np.ndarray,
    congestion: np.ndarray | None = None,
    incidents: np.ndarray | None = None,
    environmental_impact: np.ndarray | None = None,
    humidity: np.ndarray | None = None,
) -> np.ndarray:
    traffic_pressure = _minmax_scale(traffic_flow)
    congestion_pressure = _minmax_scale(congestion) if congestion is not None else np.zeros_like(traffic_pressure)
    incident_pressure = _minmax_scale(incidents) if incidents is not None else np.zeros_like(traffic_pressure)
    impact_pressure = _minmax_scale(environmental_impact) if environmental_impact is not None else np.zeros_like(traffic_pressure)
    humidity_pressure = _minmax_scale(humidity) if humidity is not None else np.zeros_like(traffic_pressure)

    traffic_dominant_signal = (
        0.72 * traffic_pressure
        + 0.18 * congestion_pressure
        + 0.05 * incident_pressure
        + 0.03 * impact_pressure
        + 0.02 * humidity_pressure
    )
    derived_aqi = 22 + traffic_dominant_signal * 248
    return (0.15 * baseline_aqi + 0.85 * derived_aqi).clip(0, 500)


def _date_from_probe_filename(path: Path) -> pd.Timestamp:
    match = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if not match:
        raise ValueError(f"Cannot infer date from Delhi probe file name: {path}")
    return pd.to_datetime(match.group(1))


def _load_delhi_probe_counts(probe_dir: Path) -> pd.DataFrame:
    files = sorted(probe_dir.glob("new_delhi__*_to_*_.geojson"))
    if not files:
        raise FileNotFoundError(f"No Delhi probe-count GeoJSON files found under {probe_dir}")

    rows = []
    for path in files:
        day = _date_from_probe_filename(path)
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)

        hourly_counts = np.zeros(24, dtype=float)
        speed_limits = []
        distances = []
        frc_values = []
        for feature in payload.get("features", []):
            props = feature.get("properties", {})
            counts = props.get("segmentProbeCounts")
            if not counts:
                continue
            speed_limits.append(float(props.get("speedLimit") or 0))
            distances.append(float(props.get("distance") or 0))
            frc_values.append(float(props.get("frc") or 0))
            for item in counts:
                time_set = int(item.get("timeSet", 0))
                hour = time_set - 2
                if 0 <= hour < 24:
                    hourly_counts[hour] += float(item.get("probeCount") or 0)

        speed_limit = float(np.mean(speed_limits)) if speed_limits else 35.0
        network_distance = float(np.sum(distances)) if distances else 1.0
        frc_mean = float(np.mean(frc_values)) if frc_values else 5.0
        for hour, probe_count in enumerate(hourly_counts):
            timestamp = day + pd.Timedelta(hours=hour)
            traffic_flow = probe_count
            congestion = np.clip((traffic_flow / max(network_distance, 1.0)) * 120 + (6 - frc_mean) * 4, 0, 100)
            rows.append(
                {
                    "timestamp": timestamp,
                    "traffic_flow": traffic_flow,
                    "traffic_congestion": congestion,
                    "traffic_speed": np.clip(speed_limit * (1.0 - congestion / 145.0), 8.0, speed_limit),
                    "traffic_incidents": max(0.0, congestion - 55.0) * 0.08,
                    "traffic_environmental_impact": congestion * traffic_flow / max(hourly_counts.max(), 1.0),
                }
            )

    return pd.DataFrame(rows).sort_values("timestamp").reset_index(drop=True)


def _expand_daily_delhi_signal(daily_df: pd.DataFrame, timestamp_col: str, value_columns: list[str]) -> pd.DataFrame:
    daily = daily_df.copy()
    daily["timestamp"] = pd.to_datetime(daily[timestamp_col], errors="coerce")
    daily = daily.dropna(subset=["timestamp"]).sort_values("timestamp")
    frames = []
    for _, row in daily.iterrows():
        hours = pd.date_range(row["timestamp"].normalize(), periods=24, freq="h")
        frame = pd.DataFrame({"timestamp": hours})
        for col in value_columns:
            frame[col] = pd.to_numeric(row[col], errors="coerce")
        frames.append(frame)
    return pd.concat(frames, ignore_index=True).drop_duplicates("timestamp", keep="last")


def _align_hourly_values(source_df: pd.DataFrame, target_timestamps: pd.Series, value_columns: list[str]) -> pd.DataFrame:
    source = source_df.copy()
    source["timestamp"] = pd.to_datetime(source["timestamp"], errors="coerce")
    source = source.dropna(subset=["timestamp"]).sort_values("timestamp").drop_duplicates(subset=["timestamp"], keep="last")
    source = source.set_index("timestamp")

    target_index = pd.DatetimeIndex(pd.to_datetime(target_timestamps))
    aligned = source.reindex(target_index)
    aligned[value_columns] = aligned[value_columns].apply(pd.to_numeric, errors="coerce")
    aligned[value_columns] = aligned[value_columns].interpolate(method="time", limit_direction="both")
    aligned[value_columns] = aligned[value_columns].ffill().bfill()
    return aligned[value_columns].reset_index(drop=True)


def _normalize_excel_header(value) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    return text


def _parse_bangalore_aqi_workbook(path: Path) -> pd.DataFrame:
    import xlrd

    workbook = xlrd.open_workbook(path)
    rows = []
    for sheet_name in workbook.sheet_names():
        sheet = workbook.sheet_by_name(sheet_name)
        header_row = None
        header_values = None
        for row_idx in range(min(sheet.nrows, 8)):
            values = [_normalize_excel_header(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
            if not values:
                continue
            first_value = values[0].lower()
            if first_value.startswith("date") and any("aqi" in value.lower() for value in values):
                header_row = row_idx
                header_values = values
                break
        if header_row is None or header_values is None:
            continue

        date_idx = 0
        aqi_idx = next((idx for idx, value in enumerate(header_values) if "aqi" in value.lower()), None)
        temp_idx = next((idx for idx, value in enumerate(header_values) if "temp" in value.lower()), None)
        humidity_idx = next(
            (
                idx
                for idx, value in enumerate(header_values)
                if "rh" in value.lower() or value.lower().startswith("hr")
            ),
            None,
        )
        if aqi_idx is None:
            continue

        for row_idx in range(header_row + 1, sheet.nrows):
            raw_date = sheet.cell_value(row_idx, date_idx)
            date_text = _normalize_excel_header(raw_date)
            if not date_text:
                continue
            timestamp = pd.to_datetime(date_text, dayfirst=True, errors="coerce")
            if pd.isna(timestamp):
                continue

            record = {"timestamp": timestamp.normalize()}
            record["aqi"] = pd.to_numeric(str(sheet.cell_value(row_idx, aqi_idx)).replace("*", "").replace("-", ""), errors="coerce")
            if temp_idx is not None:
                record["temperature"] = pd.to_numeric(
                    str(sheet.cell_value(row_idx, temp_idx)).replace("*", "").replace("-", ""),
                    errors="coerce",
                )
            if humidity_idx is not None:
                record["humidity"] = pd.to_numeric(
                    str(sheet.cell_value(row_idx, humidity_idx)).replace("*", "").replace("-", ""),
                    errors="coerce",
                )
            rows.append(record)

    if not rows:
        raise ValueError(f"No AQI rows parsed from Bangalore workbook: {path}")

    frame = pd.DataFrame(rows)
    aggregations = {"aqi": "median"}
    if "temperature" in frame.columns:
        aggregations["temperature"] = "mean"
    if "humidity" in frame.columns:
        aggregations["humidity"] = "mean"
    return frame.groupby("timestamp", as_index=False).agg(aggregations)


def _load_bangalore_aqi_daily(dataset_dir: Path) -> pd.DataFrame:
    workbook_paths = sorted(dataset_dir.glob("*AQI*.xls")) + sorted(dataset_dir.glob("*data for Bengaluru*.xls"))
    if not workbook_paths:
        raise FileNotFoundError(f"No Bangalore AQI workbooks found under {dataset_dir}")

    frames = [_parse_bangalore_aqi_workbook(path) for path in workbook_paths]
    merged = pd.concat(frames, ignore_index=True)
    aggregations = {"aqi": "median"}
    if "temperature" in merged.columns:
        aggregations["temperature"] = "mean"
    if "humidity" in merged.columns:
        aggregations["humidity"] = "mean"
    return merged.groupby("timestamp", as_index=False).agg(aggregations).sort_values("timestamp").reset_index(drop=True)


def _load_bangalore_electricity_hourly(power_dir: Path) -> pd.DataFrame:
    from openpyxl import load_workbook

    workbook_paths = sorted(power_dir.glob("ALLOCATIONVSACTUAL*.xlsx"))
    if not workbook_paths:
        raise FileNotFoundError(f"No Bangalore BESCOM load-curve workbooks found under {power_dir}")

    rows = []
    for path in workbook_paths:
        day_match = re.search(r"(\d{2}-\d{2}-\d{4})", path.name)
        if not day_match:
            continue
        day = pd.to_datetime(day_match.group(1), dayfirst=True, errors="coerce")
        if pd.isna(day):
            continue

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        sheet_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if len(sheet_rows) < 4:
            continue

        total_row = None
        for row in sheet_rows:
            first_value = row[0] if len(row) > 0 else None
            if str(first_value).strip().upper() == "BESCOM TOTAL":
                total_row = row
                break
        if total_row is None:
            continue

        time_row = sheet_rows[2]
        label_row = sheet_rows[3]
        for col_idx, label in enumerate(label_row):
            if str(label).strip().lower() != "actuals":
                continue
            hour_source = time_row[col_idx] if col_idx < len(time_row) else None
            if pd.isna(hour_source) and col_idx > 0:
                hour_source = time_row[col_idx - 1]
            hour_value = pd.to_datetime(str(hour_source), errors="coerce")
            if pd.isna(hour_value):
                continue
            actual_source = total_row[col_idx] if col_idx < len(total_row) else None
            actual_value = pd.to_numeric(actual_source, errors="coerce")
            if pd.isna(actual_value):
                continue
            rows.append(
                {
                    "timestamp": day.normalize() + pd.Timedelta(hours=int(hour_value.hour)),
                    "electricity_demand": float(actual_value),
                }
            )

    if not rows:
        raise ValueError(f"No hourly Bangalore electricity rows parsed from {power_dir}")
    return pd.DataFrame(rows).sort_values("timestamp").drop_duplicates(subset=["timestamp"], keep="last").reset_index(drop=True)


def build_bangalore_multidomain_dataset(config) -> pd.DataFrame:
    dataset_dir = Path(config.dataset_dir)
    traffic_path = dataset_dir / "Banglore_traffic_Dataset.csv"
    weather_path = dataset_dir / "export.csv"
    power_dir = dataset_dir / "BESCOM_2024_LoadCurves"

    required = [traffic_path, weather_path, power_dir]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required Bangalore dataset files: {missing}")

    traffic_raw = pd.read_csv(traffic_path)
    traffic_daily = (
        traffic_raw.groupby("Date", as_index=False)
        .agg(
            traffic_flow=("Traffic Volume", "sum"),
            traffic_speed=("Average Speed", "mean"),
            traffic_congestion=("Congestion Level", "mean"),
            traffic_incidents=("Incident Reports", "sum"),
            traffic_environmental_impact=("Environmental Impact", "mean"),
        )
        .rename(columns={"Date": "timestamp"})
    )
    traffic_daily["timestamp"] = pd.to_datetime(traffic_daily["timestamp"], errors="coerce")
    traffic_daily = traffic_daily.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
    traffic_daily = traffic_daily[traffic_daily["timestamp"].dt.year == 2024].reset_index(drop=True)

    weather_daily = pd.read_csv(weather_path)
    weather_daily["timestamp"] = pd.to_datetime(weather_daily["date"], errors="coerce")
    weather_daily = weather_daily[["timestamp", "tavg"]].rename(columns={"tavg": "weather_temperature"})
    weather_daily["weather_temperature"] = pd.to_numeric(weather_daily["weather_temperature"], errors="coerce")
    weather_daily = weather_daily.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)

    aqi_daily = _load_bangalore_aqi_daily(dataset_dir)
    electricity_hourly = _load_bangalore_electricity_hourly(power_dir)

    overlap_start = max(
        traffic_daily["timestamp"].min(),
        weather_daily["timestamp"].min(),
        aqi_daily["timestamp"].min(),
        electricity_hourly["timestamp"].min().normalize(),
    )
    overlap_end = min(
        traffic_daily["timestamp"].max(),
        weather_daily["timestamp"].max(),
        aqi_daily["timestamp"].max(),
        electricity_hourly["timestamp"].max().normalize(),
    )
    if overlap_start > overlap_end:
        raise ValueError("No overlapping Bangalore date range found across traffic, weather, AQI, and electricity sources.")

    traffic_daily = traffic_daily[
        traffic_daily["timestamp"].between(overlap_start, overlap_end)
    ].reset_index(drop=True)
    traffic_hourly = _expand_daily_traffic_to_hourly(traffic_daily)
    target_timestamps = traffic_hourly["timestamp"]

    daily_environment = weather_daily.merge(aqi_daily, on="timestamp", how="outer")
    daily_environment["temperature"] = daily_environment[["weather_temperature", "temperature"]].mean(axis=1)
    daily_environment = daily_environment[["timestamp", "aqi", "temperature", "humidity"]].sort_values("timestamp")
    environment_hourly = _expand_daily_delhi_signal(daily_environment, "timestamp", ["aqi", "temperature", "humidity"])

    aligned_environment = _align_hourly_values(environment_hourly, target_timestamps, ["aqi", "temperature", "humidity"])
    aligned_electricity = _align_hourly_values(electricity_hourly, target_timestamps, ["electricity_demand"])

    merged = pd.concat(
        [
            traffic_hourly[
                ["timestamp", "traffic_flow", "traffic_congestion", "traffic_incidents", "traffic_environmental_impact"]
            ].reset_index(drop=True),
            aligned_environment,
            aligned_electricity,
        ],
        axis=1,
    )
    return merged[["timestamp", "traffic_flow", "aqi", "electricity_demand", "temperature", "humidity"]]


def build_delhi_multidomain_dataset(config) -> pd.DataFrame:
    dataset_dir = Path(config.dataset_dir)
    traffic_probe_dir = dataset_dir / "new_delhi_traffic_dataset" / "probe_counts" / "geojson"
    weather_path = dataset_dir / "kaggel_weather_2013_to_2024.csv"
    electricity_path = dataset_dir / "electricity" / "df_final.csv"

    required = [traffic_probe_dir, weather_path, electricity_path]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required Delhi dataset files: {missing}")

    traffic_hourly = _load_delhi_probe_counts(traffic_probe_dir)
    target_timestamps = traffic_hourly["timestamp"]
    target_len = len(traffic_hourly)

    weather_raw = pd.read_csv(weather_path)
    weather_daily = weather_raw[["DATE", "temp", "humidity"]].copy()
    weather_daily[["temp", "humidity"]] = weather_daily[["temp", "humidity"]].apply(pd.to_numeric, errors="coerce")
    weather_hourly = _expand_daily_delhi_signal(weather_daily, "DATE", ["temp", "humidity"]).rename(
        columns={"temp": "temperature"}
    )
    weather_calendar = _calendar_align(
        weather_hourly,
        target_timestamps,
        ["temperature", "humidity"],
        grouping_levels=[
            ("month", "day_of_week", "hour"),
            ("month", "hour"),
            ("day_of_week", "hour"),
            ("hour",),
        ],
    )
    weather_aligned = pd.DataFrame(
        {
            "temperature": _blend_aligned_signals(
                weather_calendar["temperature"].to_numpy(),
                _resample_sequence(weather_hourly["temperature"].to_numpy(), target_len),
                calendar_weight=0.85,
            ),
            "humidity": _blend_aligned_signals(
                weather_calendar["humidity"].to_numpy(),
                _resample_sequence(weather_hourly["humidity"].to_numpy(), target_len),
                calendar_weight=0.85,
            ),
        }
    )

    electricity_raw = pd.read_csv(electricity_path)
    electricity_daily = electricity_raw[["DATE", "POWER_DEMAND"]].copy()
    electricity_daily["POWER_DEMAND"] = pd.to_numeric(electricity_daily["POWER_DEMAND"], errors="coerce")
    electricity_hourly = _expand_daily_delhi_signal(electricity_daily, "DATE", ["POWER_DEMAND"])
    electricity_calendar = _calendar_align(
        electricity_hourly.rename(columns={"POWER_DEMAND": "electricity_demand"}),
        target_timestamps,
        ["electricity_demand"],
        grouping_levels=[
            ("month", "day_of_week", "hour"),
            ("month", "hour"),
            ("day_of_week", "hour"),
            ("hour",),
        ],
    )

    hour = target_timestamps.dt.hour.to_numpy()
    demand_shape = 1.0 + 0.06 * np.maximum(0, np.sin(2 * np.pi * (hour - 8) / 24))
    electricity_demand = _blend_aligned_signals(
        electricity_calendar["electricity_demand"].to_numpy() * demand_shape,
        _resample_sequence(electricity_hourly["POWER_DEMAND"].to_numpy(), target_len),
        calendar_weight=0.75,
    )

    merged = pd.concat(
        [
            traffic_hourly[
                ["timestamp", "traffic_flow", "traffic_congestion", "traffic_incidents", "traffic_environmental_impact"]
            ].reset_index(drop=True),
            weather_aligned,
        ],
        axis=1,
    )
    baseline_aqi = 72 + 1.45 * _minmax_scale(merged["traffic_flow"].to_numpy()) * 100
    merged["aqi"] = _compute_traffic_sensitive_aqi(
        baseline_aqi=baseline_aqi,
        traffic_flow=merged["traffic_flow"].to_numpy(),
        congestion=merged["traffic_congestion"].to_numpy(),
        incidents=merged["traffic_incidents"].to_numpy(),
        environmental_impact=merged["traffic_environmental_impact"].to_numpy(),
        humidity=merged["humidity"].to_numpy(),
    )
    merged["electricity_demand"] = electricity_demand
    return merged[["timestamp", "traffic_flow", "aqi", "electricity_demand", "temperature", "humidity"]]


def build_multidomain_dataset_from_bundle(config) -> pd.DataFrame:
    dataset_dir = Path(config.dataset_dir)
    traffic_path = dataset_dir / "Banglore_traffic_Dataset.csv"
    aqi_path = dataset_dir / "AQI Data.csv"
    weather_path = dataset_dir / "temparature.csv"
    electricity_path = dataset_dir / "electricity data.csv"

    required = [traffic_path, aqi_path, weather_path, electricity_path]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing required dataset files: {missing}")

    traffic_raw = pd.read_csv(traffic_path)
    traffic_daily = (
        traffic_raw.groupby("Date", as_index=False)
        .agg(
            traffic_flow=("Traffic Volume", "sum"),
            traffic_speed=("Average Speed", "mean"),
            traffic_congestion=("Congestion Level", "mean"),
            traffic_incidents=("Incident Reports", "sum"),
            traffic_environmental_impact=("Environmental Impact", "mean"),
        )
        .rename(columns={"Date": "timestamp"})
    )
    traffic_daily["timestamp"] = pd.to_datetime(traffic_daily["timestamp"])
    traffic_hourly = _expand_daily_traffic_to_hourly(traffic_daily)

    weather_raw = pd.read_csv(weather_path)
    weather_raw["timestamp"] = pd.to_datetime(weather_raw["date_time"], errors="coerce")
    weather_hourly = (
        weather_raw[["timestamp", "tempC", "humidity"]]
        .dropna(subset=["timestamp"])
        .sort_values("timestamp")
        .drop_duplicates(subset=["timestamp"], keep="last")
        .reset_index(drop=True)
        .rename(columns={"tempC": "temperature"})
    )
    weather_hourly[["temperature", "humidity"]] = weather_hourly[["temperature", "humidity"]].apply(
        pd.to_numeric, errors="coerce"
    )
    weather_hourly[["temperature", "humidity"]] = weather_hourly[["temperature", "humidity"]].interpolate(
        method="linear", limit_direction="both"
    )

    aqi_raw = pd.read_csv(aqi_path)
    aqi_daily = aqi_raw.dropna(how="all").reset_index(drop=True)
    aqi_daily["aqi"] = pd.to_numeric(aqi_daily["PM 2.5"], errors="coerce")
    aqi_daily["humidity"] = pd.to_numeric(aqi_daily["H"], errors="coerce")
    aqi_daily["temperature"] = pd.to_numeric(aqi_daily["T"], errors="coerce")
    aqi_daily = aqi_daily[["aqi", "temperature", "humidity"]].interpolate(method="linear", limit_direction="both")
    aqi_daily["date"] = pd.date_range(traffic_daily["timestamp"].min(), periods=len(aqi_daily), freq="D")
    aqi_hourly = (
        aqi_daily.set_index("date")[["aqi"]]
        .reindex(pd.date_range(aqi_daily["date"].min(), aqi_daily["date"].max(), freq="h"))
        .interpolate(method="linear", limit_direction="both")
        .rename_axis("timestamp")
        .reset_index()
    )

    electricity_raw = pd.read_csv(electricity_path)
    electricity_yearly = electricity_raw[["Year", "Energy Input in MU"]].copy()
    electricity_yearly["start_year"] = electricity_yearly["Year"].astype(str).str.split("-").str[0].astype(int)
    electricity_yearly["timestamp"] = pd.to_datetime(electricity_yearly["start_year"].astype(str) + "-04-01")
    electricity_yearly = electricity_yearly.sort_values("timestamp")

    target_len = len(traffic_hourly)
    traffic_timeline = traffic_hourly["timestamp"].reset_index(drop=True)

    weather_calendar = _calendar_align(
        weather_hourly,
        traffic_timeline,
        ["temperature", "humidity"],
        grouping_levels=[
            ("month", "day_of_week", "hour"),
            ("month", "hour"),
            ("day_of_week", "hour"),
            ("hour",),
        ],
    )
    weather_aligned = pd.DataFrame(
        {
            "temperature_weather": _blend_aligned_signals(
                weather_calendar["temperature"].to_numpy(),
                _resample_sequence(weather_hourly["temperature"].to_numpy(), target_len),
                calendar_weight=0.7,
            ),
            "humidity_weather": _blend_aligned_signals(
                weather_calendar["humidity"].to_numpy(),
                _resample_sequence(weather_hourly["humidity"].to_numpy(), target_len),
                calendar_weight=0.7,
            ),
        }
    )
    aqi_calendar = _calendar_align(
        aqi_hourly,
        traffic_timeline,
        ["aqi"],
        grouping_levels=[
            ("month", "day_of_week", "hour"),
            ("month", "hour"),
            ("day_of_week", "hour"),
            ("hour",),
        ],
    )
    aqi_aligned = pd.DataFrame(
        {
            "aqi_baseline": _blend_aligned_signals(
                aqi_calendar["aqi"].to_numpy(),
                _resample_sequence(aqi_hourly["aqi"].to_numpy(), target_len),
                calendar_weight=0.6,
            )
        }
    )

    merged = pd.concat(
        [
            traffic_timeline,
            traffic_hourly[
                ["traffic_flow", "traffic_congestion", "traffic_incidents", "traffic_environmental_impact"]
            ].reset_index(drop=True),
            aqi_aligned,
            weather_aligned,
        ],
        axis=1,
    )

    merged["aqi"] = _compute_traffic_sensitive_aqi(
        baseline_aqi=merged["aqi_baseline"].to_numpy(),
        traffic_flow=merged["traffic_flow"].to_numpy(),
        congestion=merged["traffic_congestion"].to_numpy(),
        incidents=merged["traffic_incidents"].to_numpy(),
        environmental_impact=merged["traffic_environmental_impact"].to_numpy(),
        humidity=merged["humidity_weather"].to_numpy(),
    )

    hour = merged["timestamp"].dt.hour.to_numpy()
    weekday = merged["timestamp"].dt.dayofweek.to_numpy()
    weekend_factor = np.where(weekday >= 5, 0.95, 1.0)
    peak_cycle = 8 * np.maximum(0, np.sin(2 * np.pi * (hour - 8) / 24))
    merged["electricity_demand"] = _build_electricity_baseline(
        electricity_yearly,
        merged["timestamp"],
        merged["traffic_flow"].to_numpy(),
        merged["temperature_weather"].to_numpy(),
        merged["humidity_weather"].to_numpy(),
    ) * weekend_factor + peak_cycle

    merged = merged.rename(
        columns={
            "temperature_weather": "temperature",
            "humidity_weather": "humidity",
        }
    )
    return merged[["timestamp", "traffic_flow", "aqi", "electricity_demand", "temperature", "humidity"]]


def build_multidomain_dataset_from_raw_sources(config) -> pd.DataFrame:
    traffic_path = Path(config.raw_traffic_file)
    aqi_path = Path(config.raw_aqi_file)
    if not traffic_path.exists() or not aqi_path.exists():
        raise FileNotFoundError("Raw traffic and AQI source files are required to build the merged dataset")

    traffic_raw = pd.read_csv(traffic_path)
    traffic_daily = (
        traffic_raw.groupby("Date", as_index=False)
        .agg(
            traffic_flow=("Traffic Volume", "sum"),
            traffic_speed=("Average Speed", "mean"),
            traffic_congestion=("Congestion Level", "mean"),
            traffic_incidents=("Incident Reports", "sum"),
            traffic_environmental_impact=("Environmental Impact", "mean"),
        )
        .rename(columns={"Date": "timestamp"})
    )
    traffic_daily["timestamp"] = pd.to_datetime(traffic_daily["timestamp"])
    traffic_hourly = _expand_daily_traffic_to_hourly(traffic_daily)

    aqi_raw = pd.read_csv(aqi_path, sep=";", decimal=",")
    aqi_raw = aqi_raw.loc[:, ~aqi_raw.columns.astype(str).str.startswith("Unnamed")]
    aqi_raw = aqi_raw.drop(columns=[col for col in aqi_raw.columns if not str(col).strip()], errors="ignore")
    aqi_raw["timestamp"] = pd.to_datetime(
        aqi_raw["Date"].astype(str) + " " + aqi_raw["Time"].astype(str).str.replace(".", ":", regex=False),
        dayfirst=True,
        errors="coerce",
    )

    numeric_columns = ["CO(GT)", "NO2(GT)", "NOx(GT)", "C6H6(GT)", "T", "RH", "AH"]
    for col in numeric_columns:
        aqi_raw[col] = _to_numeric_series(aqi_raw[col])

    aqi_hourly = (
        aqi_raw[["timestamp", *numeric_columns]]
        .dropna(subset=["timestamp"])
        .sort_values("timestamp")
        .drop_duplicates(subset=["timestamp"], keep="last")
        .reset_index(drop=True)
    )
    aqi_hourly[numeric_columns] = aqi_hourly[numeric_columns].interpolate(method="linear", limit_direction="both")

    aqi_components = aqi_hourly[["CO(GT)", "NO2(GT)", "NOx(GT)", "C6H6(GT)"]]
    component_min = aqi_components.min()
    component_range = (aqi_components.max() - component_min).replace(0, 1)
    normalized = (aqi_components - component_min) / component_range
    aqi_hourly["aqi"] = (normalized.mean(axis=1) * 400 + 50).clip(0, 500)
    aqi_hourly["temperature"] = aqi_hourly["T"]
    aqi_hourly["humidity"] = aqi_hourly["RH"]

    target_len = len(traffic_hourly)
    aligned_aqi = pd.DataFrame(
        {
            "aqi_baseline": _resample_sequence(aqi_hourly["aqi"].to_numpy(), target_len),
            "temperature": _resample_sequence(aqi_hourly["temperature"].to_numpy(), target_len),
            "humidity": _resample_sequence(aqi_hourly["humidity"].to_numpy(), target_len),
        }
    )

    merged = pd.concat(
        [
            traffic_hourly[
                ["timestamp", "traffic_flow", "traffic_congestion", "traffic_incidents", "traffic_environmental_impact"]
            ].reset_index(drop=True),
            aligned_aqi,
        ],
        axis=1,
    )
    merged["aqi"] = _compute_traffic_sensitive_aqi(
        baseline_aqi=merged["aqi_baseline"].to_numpy(),
        traffic_flow=merged["traffic_flow"].to_numpy(),
        congestion=merged["traffic_congestion"].to_numpy(),
        incidents=merged["traffic_incidents"].to_numpy(),
        environmental_impact=merged["traffic_environmental_impact"].to_numpy(),
        humidity=merged["humidity"].to_numpy(),
    )
    hour = merged["timestamp"].dt.hour.to_numpy()
    weekday = merged["timestamp"].dt.dayofweek.to_numpy()
    weekend_factor = np.where(weekday >= 5, 0.94, 1.0)
    demand_signal = (
        180
        + 0.00011 * merged["traffic_flow"].to_numpy()
        + 2.6 * merged["temperature"].to_numpy()
        + 0.12 * merged["humidity"].to_numpy()
        + 14 * np.maximum(0, np.sin(2 * np.pi * (hour - 8) / 24))
    )
    merged["electricity_demand"] = demand_signal * weekend_factor
    return merged[["timestamp", "traffic_flow", "aqi", "electricity_demand", "temperature", "humidity"]]


def load_input_dataframe(config) -> pd.DataFrame:
    data_path = Path(config.data_file)
    if getattr(config, "city", "default") == "delhi":
        source_files = [
            *sorted((Path(config.dataset_dir) / "new_delhi_traffic_dataset" / "probe_counts" / "geojson").glob("*.geojson")),
            Path(config.dataset_dir) / "kaggel_weather_2013_to_2024.csv",
            Path(config.dataset_dir) / "electricity" / "df_final.csv",
        ]
        if _needs_rebuild(data_path, source_files):
            df = build_delhi_multidomain_dataset(config)
            data_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(data_path, index=False)
        else:
            df = pd.read_csv(data_path)
        df["timestamp"] = pd.to_datetime(df["timestamp"])
        expected_columns = {"timestamp", *config.domain_columns}
        missing_columns = expected_columns.difference(df.columns)
        if missing_columns:
            raise ValueError(f"Missing required columns in dataset: {sorted(missing_columns)}")
        df = df.sort_values("timestamp").reset_index(drop=True)
        return clean_fuzzy_data(df, config)

    if getattr(config, "city", "default") == "bangalore":
        aqi_workbooks = sorted(Path(config.dataset_dir).glob("*AQI*.xls")) + sorted(
            Path(config.dataset_dir).glob("*data for Bengaluru*.xls")
        )
        source_files = [
            Path(config.dataset_dir) / "Banglore_traffic_Dataset.csv",
            Path(config.dataset_dir) / "export.csv",
            *aqi_workbooks,
            *sorted((Path(config.dataset_dir) / "BESCOM_2024_LoadCurves").glob("*.xlsx")),
        ]
        if source_files and _needs_rebuild(data_path, source_files):
            df = build_bangalore_multidomain_dataset(config)
            data_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(data_path, index=False)
        elif data_path.exists():
            df = pd.read_csv(data_path)
        else:
            df = build_bangalore_multidomain_dataset(config)
            data_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(data_path, index=False)

        df["timestamp"] = pd.to_datetime(df["timestamp"])
        expected_columns = {"timestamp", *config.domain_columns}
        missing_columns = expected_columns.difference(df.columns)
        if missing_columns:
            raise ValueError(f"Missing required columns in dataset: {sorted(missing_columns)}")
        df = df.sort_values("timestamp").reset_index(drop=True)
        return clean_fuzzy_data(df, config)

    bundle_files = [
        Path(config.dataset_dir) / "Banglore_traffic_Dataset.csv",
        Path(config.dataset_dir) / "AQI Data.csv",
        Path(config.dataset_dir) / "temparature.csv",
        Path(config.dataset_dir) / "electricity data.csv",
    ]
    if all(path.exists() for path in bundle_files):
        if _needs_rebuild(data_path, bundle_files):
            df = build_multidomain_dataset_from_bundle(config)
            data_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(data_path, index=False)
        else:
            df = pd.read_csv(data_path)
    elif data_path.exists():
        df = pd.read_csv(data_path)
    elif Path(config.raw_traffic_file).exists() and Path(config.raw_aqi_file).exists():
        raw_files = [Path(config.raw_traffic_file), Path(config.raw_aqi_file)]
        if _needs_rebuild(data_path, raw_files):
            df = build_multidomain_dataset_from_raw_sources(config)
            data_path.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(data_path, index=False)
        else:
            df = pd.read_csv(data_path)
    else:
        df = generate_synthetic_multidomain_data(config.num_steps, config.random_seed)

    df["timestamp"] = pd.to_datetime(df["timestamp"])
    expected_columns = {"timestamp", *config.domain_columns}
    missing_columns = expected_columns.difference(df.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns in dataset: {sorted(missing_columns)}")
    df = df.sort_values("timestamp").reset_index(drop=True)
    return clean_fuzzy_data(df, config)


def clean_fuzzy_data(df: pd.DataFrame, config) -> pd.DataFrame:
    cleaned = df.copy()
    numeric_columns = [col for col in config.domain_columns if col in cleaned.columns]

    # Remove duplicate timestamps and preserve the latest reading.
    cleaned = cleaned.drop_duplicates(subset=["timestamp"], keep="last").reset_index(drop=True)

    # Convert bad strings or corrupt values into missing values first.
    for col in numeric_columns:
        cleaned[col] = pd.to_numeric(cleaned[col], errors="coerce")

    # Use timestamp-aware interpolation so missing values follow the time axis rather than row position only.
    timestamp_indexed = cleaned.set_index("timestamp")
    timestamp_indexed[numeric_columns] = timestamp_indexed[numeric_columns].interpolate(
        method="time", limit_direction="both"
    )
    cleaned = timestamp_indexed.reset_index()

    # Clip extreme outliers with an IQR fence to make the training signal more stable.
    for col in numeric_columns:
        q1 = cleaned[col].quantile(0.25)
        q3 = cleaned[col].quantile(0.75)
        iqr = q3 - q1
        if iqr <= 0:
            continue
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        cleaned[col] = cleaned[col].clip(lower=lower, upper=upper)

    cleaned[numeric_columns] = cleaned[numeric_columns].ffill().bfill()
    return cleaned


def add_time_features(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    ts = pd.to_datetime(out["timestamp"])
    hour = ts.dt.hour
    day_of_week = ts.dt.dayofweek
    day_of_year = ts.dt.dayofyear

    out["hour"] = hour
    out["day_of_week"] = day_of_week
    out["month"] = ts.dt.month
    out["is_weekend"] = (day_of_week >= 5).astype(int)
    out["hour_sin"] = np.sin(2 * np.pi * hour / 24)
    out["hour_cos"] = np.cos(2 * np.pi * hour / 24)
    out["dow_sin"] = np.sin(2 * np.pi * day_of_week / 7)
    out["dow_cos"] = np.cos(2 * np.pi * day_of_week / 7)
    out["season_sin"] = np.sin(2 * np.pi * day_of_year / 365)
    out["season_cos"] = np.cos(2 * np.pi * day_of_year / 365)
    return out


def add_lag_features(df: pd.DataFrame, columns: List[str], lag_steps: Tuple[int, ...]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        for lag in lag_steps:
            out[f"{col}_lag_{lag}"] = out[col].shift(lag)
    return out


def add_rolling_features(df: pd.DataFrame, columns: List[str], windows: Tuple[int, ...]) -> pd.DataFrame:
    out = df.copy()
    for col in columns:
        for window in windows:
            roll = out[col].rolling(window=window)
            out[f"{col}_roll_mean_{window}"] = roll.mean()
            out[f"{col}_roll_std_{window}"] = roll.std()
            out[f"{col}_roll_min_{window}"] = roll.min()
            out[f"{col}_roll_max_{window}"] = roll.max()
    return out


class DataProcessor:
    def __init__(self, config):
        self.config = config
        self.scaler = StandardScaler() if config.scaler_name == "standard" else MinMaxScaler()
        self.target_scaler = StandardScaler() if config.scaler_name == "standard" else MinMaxScaler()
        self.feature_columns: List[str] = []
        self.target_columns = list(config.target_columns)

    def prepare_dataframe(self, raw_df: pd.DataFrame) -> pd.DataFrame:
        df = add_time_features(raw_df)
        df = add_lag_features(df, list(self.config.domain_columns), self.config.lag_steps)
        df = add_rolling_features(df, list(self.config.domain_columns), self.config.rolling_windows)
        df = df.dropna().reset_index(drop=True)
        return df

    def split_dataframe(self, df: pd.DataFrame):
        n = len(df)
        train_end = int(n * self.config.train_ratio)
        val_end = int(n * (self.config.train_ratio + self.config.val_ratio))
        return df.iloc[:train_end].copy(), df.iloc[train_end:val_end].copy(), df.iloc[val_end:].copy()

    def fit_transform(self, train_df: pd.DataFrame, val_df: pd.DataFrame, test_df: pd.DataFrame):
        ignore = {"timestamp"}
        self.feature_columns = [col for col in train_df.columns if col not in ignore]

        x_train = self.scaler.fit_transform(train_df[self.feature_columns])
        x_val = self.scaler.transform(val_df[self.feature_columns])
        x_test = self.scaler.transform(test_df[self.feature_columns])

        y_train = self.target_scaler.fit_transform(train_df[self.target_columns])
        y_val = self.target_scaler.transform(val_df[self.target_columns])
        y_test = self.target_scaler.transform(test_df[self.target_columns])
        return x_train, y_train, x_val, y_val, x_test, y_test

    def transform_dataframe(self, df: pd.DataFrame):
        x = self.scaler.transform(df[self.feature_columns])
        y = self.target_scaler.transform(df[self.target_columns])
        return x, y

    def inverse_transform_targets(self, y_scaled: np.ndarray) -> np.ndarray:
        y_scaled = np.asarray(y_scaled)
        if y_scaled.ndim == 1:
            y_scaled = y_scaled.reshape(1, -1)
        return self.target_scaler.inverse_transform(y_scaled)


def build_sequences(features: np.ndarray, targets: np.ndarray, seq_len: int, horizon: int):
    xs, ys = [], []
    for end_idx in range(seq_len, len(features) - horizon + 1):
        start_idx = end_idx - seq_len
        xs.append(features[start_idx:end_idx])
        ys.append(targets[end_idx + horizon - 1])
    return np.asarray(xs, dtype=np.float32), np.asarray(ys, dtype=np.float32)


def build_temporal_groups(
    features: np.ndarray,
    targets: np.ndarray,
    timestamps: pd.Series,
    horizon: int,
    closeness_lags: Tuple[int, ...],
    period_lags: Tuple[int, ...],
    trend_lags: Tuple[int, ...],
):
    all_lags = tuple(closeness_lags) + tuple(period_lags) + tuple(trend_lags)
    max_lag = max(all_lags)

    # Seasonal-naive baselines (same hour on the previous day / previous week).
    # Only past observed targets are used, so this introduces no leakage.
    daily_period = 24
    weekly_period = 24 * 7

    closeness_x, period_x, trend_x, ys, ts = [], [], [], [], []
    seasonal_daily, seasonal_weekly = [], []
    for end_idx in range(max_lag, len(features) - horizon + 1):
        closeness_x.append(np.stack([features[end_idx - lag] for lag in closeness_lags], axis=0))
        period_x.append(np.stack([features[end_idx - lag] for lag in period_lags], axis=0))
        trend_x.append(np.stack([features[end_idx - lag] for lag in trend_lags], axis=0))
        pred_idx = end_idx + horizon - 1
        ys.append(targets[pred_idx])
        ts.append(timestamps.iloc[pred_idx])
        seasonal_daily.append(targets[max(0, pred_idx - daily_period)])
        seasonal_weekly.append(targets[max(0, pred_idx - weekly_period)])

    return {
        "closeness": np.asarray(closeness_x, dtype=np.float32),
        "period": np.asarray(period_x, dtype=np.float32),
        "trend": np.asarray(trend_x, dtype=np.float32),
        "seasonal_daily": np.asarray(seasonal_daily, dtype=np.float32),
        "seasonal_weekly": np.asarray(seasonal_weekly, dtype=np.float32),
        "target": np.asarray(ys, dtype=np.float32),
        "timestamp": pd.Series(ts, name="timestamp"),
    }


def build_temporal_groups_for_inference(features: np.ndarray, timestamps: pd.Series, config):
    dummy_targets = np.zeros((len(features), len(config.target_columns)), dtype=np.float32)
    grouped = build_temporal_groups(
        features,
        dummy_targets,
        timestamps.reset_index(drop=True),
        config.forecast_horizon,
        config.closeness_lags,
        config.period_lags,
        config.trend_lags,
    )
    return {
        "closeness": grouped["closeness"],
        "period": grouped["period"],
        "trend": grouped["trend"],
        "seasonal_daily": grouped["seasonal_daily"],
        "seasonal_weekly": grouped["seasonal_weekly"],
        "timestamp": grouped["timestamp"],
    }


def create_datasets(config, raw_df: pd.DataFrame):
    processor = DataProcessor(config)
    prepared_df = processor.prepare_dataframe(raw_df)
    train_df, val_df, test_df = processor.split_dataframe(prepared_df)
    x_train, y_train, x_val, y_val, x_test, y_test = processor.fit_transform(train_df, val_df, test_df)

    train_seq = build_sequences(x_train, y_train, config.seq_len, config.forecast_horizon)
    val_seq = build_sequences(x_val, y_val, config.seq_len, config.forecast_horizon)
    test_seq = build_sequences(x_test, y_test, config.seq_len, config.forecast_horizon)
    train_tpt = build_temporal_groups(
        x_train,
        y_train,
        train_df["timestamp"].reset_index(drop=True),
        config.forecast_horizon,
        config.closeness_lags,
        config.period_lags,
        config.trend_lags,
    )
    val_tpt = build_temporal_groups(
        x_val,
        y_val,
        val_df["timestamp"].reset_index(drop=True),
        config.forecast_horizon,
        config.closeness_lags,
        config.period_lags,
        config.trend_lags,
    )
    test_tpt = build_temporal_groups(
        x_test,
        y_test,
        test_df["timestamp"].reset_index(drop=True),
        config.forecast_horizon,
        config.closeness_lags,
        config.period_lags,
        config.trend_lags,
    )

    return {
        "processor": processor,
        "prepared_df": prepared_df,
        "train_df": train_df,
        "val_df": val_df,
        "test_df": test_df,
        "train_seq": train_seq,
        "val_seq": val_seq,
        "test_seq": test_seq,
        "train_tpt": train_tpt,
        "val_tpt": val_tpt,
        "test_tpt": test_tpt,
    }


def rolling_mean(values: List[float], window: int) -> np.ndarray:
    series = pd.Series(values)
    return series.rolling(window=window, min_periods=1).mean().to_numpy()
